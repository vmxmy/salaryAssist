# -*- coding: utf-8 -*-
"""
财政正编工资报表处理全流程脚本

包含模块：
1. 字段映射加载
2. 字段转换
3. 扣款项合并
4. 数据过滤和起始行检测
5. 合并处理流程
6. 格式化输出（样式、美化、冻结行列）

适配源数据 + 扣款数据 + 示例模板
"""

import pandas as pd
import numpy as np
import re
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# --- 1. 字段映射加载 ---
def get_identity_mapping_rules(identity_value: str, field_mappings: list, rule_identity_key: str) -> dict:
    """
    根据给定的值和规则中的键，从字段映射列表中查找匹配的规则。

    Args:
        identity_value: 从源数据中获取的用于匹配的值。
        field_mappings: 包含所有映射规则的列表。
        rule_identity_key: 在映射规则字典中用于匹配的键名。

    Returns:
        匹配的规则字典，如果未找到则返回空字典。
    """
    for rule in field_mappings:
        # 使用传入的 rule_identity_key 进行匹配
        if rule.get(rule_identity_key) == identity_value:
            # 确保返回的字典结构符合预期 (至少包含 mappings)
            return {
                "编制": rule.get("编制", ""), # 保留其他可能的键，提供默认值
                "人员身份": rule.get("人员身份", ""),
                "岗位类别": rule.get("岗位类别", ""),
                # ... 可以添加其他可能在 rule 顶层存在的键 ...
                rule_identity_key: identity_value, # 确保匹配上的键值对在结果中
                "mappings": rule.get("mappings", []) # 必须有 mappings
            }
    # print(f"DEBUG: No rule found for key '{rule_identity_key}' with value '{identity_value}'")
    return {}

# --- 2. 字段转换 ---
def apply_field_mapping(df: pd.DataFrame, mapping_rules: dict) -> pd.DataFrame:
    # print(f"DEBUG: apply_field_mapping called...") # Reduced verbosity
    result_df = pd.DataFrame()
    mappings = mapping_rules.get("mappings", [])

    # --- 新增：收集所有复杂计算需要的源字段 --- #
    complex_source_fields = set()
    for mapping in mappings:
        if "source_fields" in mapping:
            complex_source_fields.update(mapping["source_fields"])
    print(f"DEBUG: apply_field_mapping - Complex source fields needed: {complex_source_fields}")
    # --- 结束新增 --- #

    for mapping in mappings:
        target = mapping.get("target_field")
        if target is None: continue

        if "source_field" in mapping:
            source = mapping["source_field"]
            if source in df.columns:
                 result_df[target] = df[source]
            else:
                 print(f"Warning: Simple mapping source field '{source}' not found in input df for target '{target}'. Setting target to NaN.")
                 result_df[target] = np.nan # Set target to NaN if source is missing
        elif "source_fields" in mapping:
            # Defer calculation, but ensure target column might be created later if needed.
            # No action needed here now, source fields will be preserved below.
            pass

    # Add static fields from the rule (e.g., '编制', '人员身份')
    for key, value in mapping_rules.items():
        if key not in ["mappings", "persons"] + list(complex_source_fields): # Avoid overwriting complex sources if rule has same key
             if key not in result_df.columns:
                  result_df[key] = value

    # --- 新增：确保复杂计算需要的原始源字段被保留 --- #
    print(f"DEBUG: apply_field_mapping - Checking preservation for {len(complex_source_fields)} complex source fields...")
    preserved_count = 0
    already_present_count = 0
    missing_from_input_count = 0
    for field_name in complex_source_fields:
        if field_name not in result_df.columns:
            if field_name in df.columns:
                result_df[field_name] = df[field_name]
                preserved_count += 1
                # print(f"DEBUG: Preserving complex source field '{field_name}' from input df.")
            else:
                 # This case should be rare if the source file is consistent, but good to log.
                 print(f"Warning: Complex source field '{field_name}' needed for later calculation was NOT found in the input df columns: {df.columns.tolist()}")
                 missing_from_input_count += 1
        else:
            already_present_count += 1
            # print(f"DEBUG: Complex source field '{field_name}' already present in result_df.")
    print(f"DEBUG: apply_field_mapping - Preservation check complete. Preserved: {preserved_count}, Already Present: {already_present_count}, Missing from Input: {missing_from_input_count}")
    # --- 结束新增 --- #

    print(f"DEBUG: apply_field_mapping returning columns: {result_df.columns.tolist()}")
    return result_df

# --- 3. 合并扣款项 ---
def merge_deductions(source_df: pd.DataFrame, deduction_df: pd.DataFrame, deduction_fields: list) -> pd.DataFrame:
    print("DEBUG: merge_deductions called.")
    print(f"DEBUG: Source DataFrame columns: {source_df.columns.tolist()}")
    print(f"DEBUG: Source DataFrame first 5 rows:\n{source_df.head().to_string()}")
    print(f"DEBUG: Deduction DataFrame columns: {deduction_df.columns.tolist()}")
    print(f"DEBUG: Deduction DataFrame first 5 rows:\n{deduction_df.head().to_string()}")
    
    merged_df = source_df.copy()
    # --- 动态确定合并列名 (查找在两个表中都存在的列) --- #
    possible_key_columns = ["姓名", "人员姓名"] # 优先尝试 '姓名'
    merge_on_column = None
    for col in possible_key_columns:
        # Check if column exists in BOTH dataframes
        if col in merged_df.columns and col in deduction_df.columns:
            merge_on_column = col
            print(f"DEBUG: Found common merge key in both dataframes: '{col}'")
            print(f"DEBUG: Unique values in source '{col}': {merged_df[col].unique().tolist()}")
            print(f"DEBUG: Unique values in deduction '{col}': {deduction_df[col].unique().tolist()}")
            break # Found the best key

    if merge_on_column is None:
        # Check if key exists in one but not the other
        source_keys = [k for k in possible_key_columns if k in merged_df.columns]
        deduction_keys = [k for k in possible_key_columns if k in deduction_df.columns]
        print(f"ERROR: Cannot find a common merge key column. Keys in source: {source_keys}, Keys in deduction: {deduction_keys}. Skipping merge.")
        return merged_df
    # --- 结束动态确定 --- #

    # 确保源 DataFrame 包含所有需要合并的扣款字段，不存在则添加并填充 NaN
    for field in deduction_fields:
        if field not in merged_df.columns:
            print(f"DEBUG: Adding NaN column '{field}' to merged_df before merge.")
            merged_df[field] = np.nan

    # 只从扣款表中选择 merge_on_column 和指定的扣款字段进行合并
    columns_to_merge = [merge_on_column] + [field for field in deduction_fields if field in deduction_df.columns]
    print(f"DEBUG: Columns selected from deduction_df for merge: {columns_to_merge}")

    print(f"DEBUG: Performing merge on '{merge_on_column}'")
    merged_df = pd.merge(
        merged_df,
        deduction_df[columns_to_merge],
        on=merge_on_column, # <--- 使用动态确定的列名
        how="left",
        suffixes=("", "_扣款表")
    )
    print(f"DEBUG: Shape after merge: {merged_df.shape}")
    print(f"DEBUG: Merged DataFrame first 5 rows:\n{merged_df.head().to_string()}")

    # 使用扣款表中的值填充源表中的 NaN 值
    print("DEBUG: Applying np.where to prioritize deduction values...")
    for field in deduction_fields:
         deduction_col_suffixed = f"{field}_扣款表"
         if deduction_col_suffixed in merged_df.columns:
             print(f"DEBUG: Processing field '{field}'...")
             original_col = merged_df[field]
             deduction_col = merged_df[deduction_col_suffixed]
             
             print(f"DEBUG: Original column '{field}' values (first 5):\n{original_col.head().to_string()}")
             print(f"DEBUG: Deduction column '{field}' values (first 5):\n{deduction_col.head().to_string()}")

             result_col = np.where(
                 deduction_col.notna(),
                 deduction_col,
                 original_col
             )
             merged_df[field] = result_col
             
             print(f"DEBUG: Result column '{field}' values after merge (first 5):\n{merged_df[field].head().to_string()}")
             merged_df.drop(columns=[deduction_col_suffixed], inplace=True)

    print(f"DEBUG: merge_deductions returning shape: {merged_df.shape}")
    print(f"DEBUG: Final DataFrame columns: {merged_df.columns.tolist()}")
    print(f"DEBUG: Final DataFrame first 5 rows:\n{merged_df.head().to_string()}")
    return merged_df

# --- 4. 起始行检测与合计过滤 ---
def detect_data_start_row(df: pd.DataFrame, keyword: str = "人员姓名", max_scan_rows: int = 10) -> int:
    for i in range(max_scan_rows):
        row = df.iloc[i].astype(str).tolist()
        if any(keyword in str(cell) for cell in row):
            return i
    raise ValueError(f"未找到字段 '{keyword}' 所在行")

# --- 5. 批量处理函数 ---
def process_sheet(file_path: str, deduction_df: pd.DataFrame, field_mappings: list, selected_deduction_fields: list, source_identity_column: str, rule_identity_key: str) -> pd.DataFrame:
    print(f"DEBUG: process_sheet called for file: {os.path.basename(file_path)}")
    print(f"DEBUG: Using source identity column: '{source_identity_column}', rule identity key: '{rule_identity_key}'")
    try:
        preview = pd.read_excel(file_path, nrows=10, header=None)
        header_row = detect_data_start_row(preview, keyword=source_identity_column)
        print(f"DEBUG: Detected header row: {header_row}")
        df = pd.read_excel(file_path, skiprows=header_row + 1, header=None)
        df.columns = preview.iloc[header_row].tolist()
        print(f"DEBUG: Read source data, shape: {df.shape}")

        # 过滤掉合计/汇总行 (使用 source_identity_column 检查可能更可靠？取决于该列是否包含这些词)
        # 暂时保留对 人员身份 的检查，如果 source_identity_column 不同，可能需要调整
        filter_col = source_identity_column if source_identity_column in df.columns else "人员身份" # 回退到 人员身份
        if filter_col in df.columns:
            rows_before_filter = len(df)
            df = df[~df[filter_col].astype(str).str.contains("合计|汇总|总计|备注|说明", na=False)]
            print(f"DEBUG: Filtered rows based on '{filter_col}'. Shape before: {rows_before_filter}, after: {len(df)}")
        else:
            print(f"Warning: Cannot apply filter row logic as column '{filter_col}' not found.")

        results = []
        processed_ids = set()
        missing_rule_ids = set()

        print(f"DEBUG: Starting row-by-row processing using '{source_identity_column}' for identity...")
        for index, row in df.iterrows():
            identity_value = row.get(source_identity_column)
            if pd.isna(identity_value):
                continue
            identity_value = str(identity_value)
            # print(f"DEBUG: Processing row index {index}, identity_value: '{identity_value}'") # Commented out

            processed_ids.add(identity_value)
            mapping = get_identity_mapping_rules(identity_value, field_mappings, rule_identity_key)
            if not mapping:
                missing_rule_ids.add(identity_value)
                continue
            # print(f"DEBUG: Found mapping rule for '{identity_value}'")

            single_df = pd.DataFrame([row])
            converted = apply_field_mapping(single_df, mapping)
            # 添加匹配时使用的键和值到结果中，便于追溯
            converted[f'_匹配字段 ({source_identity_column})'] = identity_value
            converted[f'_匹配规则键 ({rule_identity_key})'] = mapping.get(rule_identity_key)
            results.append(converted)

        if missing_rule_ids:
             print(f"Warning: No mapping rules found for {rule_identity_key} values: {sorted(list(missing_rule_ids))}")

        if not results:
            print(f"Warning: No rows processed successfully for file {os.path.basename(file_path)}.")
            return pd.DataFrame()

        print(f"DEBUG: Concatenating {len(results)} processed rows...")
        df_combined = pd.concat(results, ignore_index=True)
        print(f"DEBUG: df_combined shape after concat: {df_combined.shape}")

        # --- 合并扣款数据 ---
        print("DEBUG: Starting deduction merge...")
        print(f"DEBUG: Selected deduction fields: {selected_deduction_fields}")
        print(f"DEBUG: Deduction DataFrame columns: {deduction_df.columns.tolist()}")
        # print(f"DEBUG: Deduction DataFrame first 5 rows:\\n{deduction_df.head().to_string()}") # Reduce log verbosity

        # 动态查找姓名列
        possible_name_cols = ["人员姓名", "姓名"]
        source_name_col = next((col for col in possible_name_cols if col in df_combined.columns), None)
        deduction_name_col = next((col for col in possible_name_cols if col in deduction_df.columns), None)

        if source_name_col and deduction_name_col:
            print(f"DEBUG: Found name column in source: '{source_name_col}'")
            print(f"DEBUG: Found name column in deduction: '{deduction_name_col}'")

            # 准备用于合并的扣款数据副本
            cols_to_merge = [deduction_name_col] + selected_deduction_fields
            missing_deduction_fields = [f for f in selected_deduction_fields if f not in deduction_df.columns]
            if missing_deduction_fields:
                 print(f"WARNING: The following selected deduction fields are missing from the deduction table: {missing_deduction_fields}")
                 cols_to_merge = [f for f in cols_to_merge if f in deduction_df.columns] # Only use existing columns

            deduction_df_to_merge = deduction_df[cols_to_merge].copy()

            # 如果姓名列名称不一致，重命名扣款表的列以匹配源表
            merge_key = source_name_col
            if source_name_col != deduction_name_col:
                print(f"DEBUG: Renaming deduction key column '{deduction_name_col}' to '{source_name_col}' for merge.")
                deduction_df_to_merge.rename(columns={deduction_name_col: source_name_col}, inplace=True)

            print(f"DEBUG: Merging on key: '{merge_key}'")
            # print(f"DEBUG: Source name values (first 5):\\n{df_combined[source_name_col].head().to_string()}")
            # print(f"DEBUG: Deduction (to merge) name values (first 5):\\n{deduction_df_to_merge[merge_key].head().to_string()}")

            # 执行合并
            df_combined = pd.merge(
                df_combined,
                deduction_df_to_merge,
                on=merge_key,
                how="left"
                # Consider adding suffixes if other column names might collide, e.g., suffixes=('', '_deduction')
            )
            print(f"DEBUG: Shape after merge: {df_combined.shape}")
            # print(f"DEBUG: Merged DataFrame columns: {df_combined.columns.tolist()}") # Reduce log verbosity
            # print(f"DEBUG: Merged DataFrame first 5 rows:\\n{df_combined.head().to_string()}") # Reduce log verbosity

            # 检查合并后是否有NaN（如果需要）
            nan_check_cols = [f for f in selected_deduction_fields if f in df_combined.columns] # Check only merged fields
            if nan_check_cols:
                 nan_counts = df_combined[nan_check_cols].isnull().sum()
                 total_nans = nan_counts.sum()
                 if total_nans > 0:
                     print(f"DEBUG: NaN values found after merge (field: count): {nan_counts[nan_counts > 0].to_dict()}")
                 else:
                     print(f"DEBUG: No NaN values found in merged deduction columns.")

        else:
            error_msg = "ERROR: Cannot perform merge. "
            if not source_name_col:
                error_msg += f"Name column ({'/'.join(possible_name_cols)}) not found in source data (df_combined columns: {df_combined.columns.tolist()}). "
            if not deduction_name_col:
                error_msg += f"Name column ({'/'.join(possible_name_cols)}) not found in deduction data (deduction_df columns: {deduction_df.columns.tolist()})."
            print(error_msg)
            # Consider adding empty columns for selected_deduction_fields if merge fails?
            # for field in selected_deduction_fields:
            #     if field not in df_combined.columns:
            #         df_combined[field] = np.nan

        # --- 应用复杂计算规则 --- #
        print(f"DEBUG: Applying complex calculations based on field mapping...")
        all_complex_mappings_details = {}
        identity_column_for_lookup = f'_匹配字段 ({source_identity_column})'
        if identity_column_for_lookup in df_combined.columns:
             unique_identity_values = df_combined[identity_column_for_lookup].unique()
             print(f"DEBUG: Found {len(unique_identity_values)} unique identity values for rule lookup from column '{identity_column_for_lookup}'")
        else:
             print(f"ERROR: Cannot find column '{identity_column_for_lookup}' in df_combined to look up rules. Skipping complex calculations.")
             unique_identity_values = []

        for identity_value in unique_identity_values:
                rule = get_identity_mapping_rules(str(identity_value), field_mappings, rule_identity_key)
                if rule:
                    for mapping in rule.get("mappings", []):
                        if "source_fields" in mapping:
                            target = mapping["target_field"]
                            if target not in all_complex_mappings_details:
                                    all_complex_mappings_details[target] = {
                                        "sources": mapping["source_fields"],
                                        "calculation": mapping.get("calculation", "sum")
                                    }

        # --- 添加日志：打印应用复杂计算前的列名 和 '补发工资' 规则细节 ---
        print(f"DEBUG: Columns available in df_combined *before* applying complex calculations: {df_combined.columns.tolist()}")
        if '补发工资' in all_complex_mappings_details:
            print(f"DEBUG: Rule details for '补发工资' from complex mappings: {all_complex_mappings_details['补发工资']}")
        else:
            print("DEBUG: No rule details found for '补发工资' in all_complex_mappings_details. It might not be a complex calculation or rule is missing.")
        # --- 结束添加 ---

        # 定义行计算辅助函数
        def calculate_row(row, sources, calculation, target):
            row_info = f"row index {row.name}" if hasattr(row, 'name') else "unknown row"
            print(f"DEBUG: calculate_row for '{target}', {row_info}")

            # Check for missing source columns *before* trying to access them
            missing_sources = [s for s in sources if s not in row.index]
            if missing_sources:
                print(f"DEBUG: Missing sources for '{target}' in {row_info}: {missing_sources}")
                return np.nan

            try:
                # Get only the source values that actually exist in the row
                existing_sources = [s for s in sources if s in row.index] # Should be all sources now if no missing_sources
                source_values = row[existing_sources]

                print(f"DEBUG: Source values for '{target}': {source_values.to_dict()}")
                print(f"DEBUG: Source values types: {source_values.apply(type)}")

                # 尝试转换为数值类型，保留原始值用于调试
                numeric_sources_nan = source_values.apply(pd.to_numeric, errors='coerce')
                print(f"DEBUG: After to_numeric(coerce): {numeric_sources_nan.to_dict()}")
                print(f"DEBUG: NaN values after conversion: {numeric_sources_nan[numeric_sources_nan.isna()].to_dict()}")

                numeric_sources = numeric_sources_nan.fillna(0)
                print(f"DEBUG: After fillna(0): {numeric_sources.to_dict()}")

                result = np.nan # Default result
                if calculation == "sum":
                    result = numeric_sources.sum(skipna=False, min_count=0)
                    print(f"DEBUG: Sum result for '{target}': {result}")

                elif isinstance(calculation, str):
                    # This part is less likely for '补发工资' if it's just a sum, but kept for completeness
                    local_vars = numeric_sources.to_dict()
                    print(f"DEBUG: Evaluating expression '{calculation}' with variables: {local_vars}")
                    result = eval(calculation, {"__builtins__": {}}, local_vars)
                    print(f"DEBUG: Eval result for '{target}': {result}")
                else:
                    print(f"Warning: Unsupported calculation type '{calculation}' for target '{target}'")

                # 验证计算结果
                if pd.isna(result):
                    print(f"DEBUG: Result is NaN for '{target}' in {row_info}")
                elif not isinstance(result, (int, float)):
                    print(f"DEBUG: Result type is {type(result)} for '{target}' in {row_info}")

                return result

            except Exception as e:
                print(f"ERROR calculating '{target}' for {row_info}:")
                print(f"  Sources attempted: {sources}")
                print(f"  Calculation: '{calculation}'")
                print(f"  Error: {str(e)}")
                print(f"  Source values (if available): {source_values.to_dict() if 'source_values' in locals() else 'N/A'}")
                return np.nan

        # 应用计算
        print(f"DEBUG: Found complex mappings for targets: {list(all_complex_mappings_details.keys())}")
        for target, details in all_complex_mappings_details.items():
                print(f"DEBUG: Calculating '{target}' using '{details['calculation']}' on {details['sources']}")
                df_combined[target] = df_combined.apply(
                    lambda row: calculate_row(row, details["sources"], details["calculation"], target),
                    axis=1
                )
                # DEBUG: 检查计算结果中 NaN 的数量
                nan_count = df_combined[target].isnull().sum()
                if nan_count > 0:
                    print(f"DEBUG: Column '{target}' calculated with {nan_count} NaN values out of {len(df_combined)}.")

        print("DEBUG: Finished applying complex calculations.")
        # --- 结束新代码块 ---

        # 计算实发工资，确保应发和扣发合计存在 (扣发合计现在由上面的复杂计算生成)
        if "应发工资" in df_combined.columns and "扣发合计" in df_combined.columns:
            try:
                # 确保应发和扣发是数值，空值填0，并强制为 float 类型
                yingfa = pd.to_numeric(df_combined["应发工资"], errors='coerce').fillna(0).astype(float)
                koufa = pd.to_numeric(df_combined["扣发合计"], errors='coerce').fillna(0).astype(float)

                # 显式处理其他补扣
                if "其他补扣" in df_combined.columns:
                    other_deductions = pd.to_numeric(df_combined["其他补扣"], errors='coerce').fillna(0).astype(float)
                else:
                    # 如果列不存在，创建一个与df_combined长度相同、值为0的Series
                    other_deductions = pd.Series(0.0, index=df_combined.index) # 使用 0.0 明确为 float

                # --- 调试信息 --- #
                print("DEBUG: Calculating 实发工资 for first 5 rows:")
                print(f"  Yingfa (dtype: {yingfa.dtype}):")
                print(yingfa.head().to_string())
                print(f"  Koufa (dtype: {koufa.dtype}):")
                print(koufa.head().to_string())
                print(f"  Other Deductions (dtype: {other_deductions.dtype}):")
                print(other_deductions.head().to_string())
                # --- 结束调试 --- #

                # 核心计算
                df_combined["实发工资"] = yingfa - koufa - other_deductions
                print("DEBUG: 实发工资 calculation potentially successful.")

            except Exception as e:
                print(f"ERROR: Exception during 实发工资 calculation (yingfa - koufa - other): {e}")
                # 计算失败时，填充 NaN
                df_combined["实发工资"] = np.nan

        elif "应发工资" in df_combined.columns:
            try:
                 # 如果没有扣发，实发=应发 (同样处理空值并强制类型)
                yingfa = pd.to_numeric(df_combined["应发工资"], errors='coerce').fillna(0).astype(float)
                print(f"DEBUG: Calculating 实发工资 (only yingfa) for first 5 rows:")
                print(f"  Yingfa (dtype: {yingfa.dtype}):")
                print(yingfa.head().to_string())
                df_combined["实发工资"] = yingfa
            except Exception as e:
                print(f"ERROR: Exception during 实发工资 calculation (only yingfa): {e}")
                df_combined["实发工资"] = np.nan
        else:
                # 如果连应发工资都没有
                print("DEBUG: Yingfa column missing, setting 实发工资 to 0.")
                df_combined["实发工资"] = 0.0 # 使用 0.0 明确为 float

        print(f"DEBUG: Finished calculating 实发工资.")
        # --- 结束实发工资计算 --- #

        print(f"DEBUG: process_sheet returning final shape: {df_combined.shape}")
        return df_combined

    except FileNotFoundError:
        print(f"ERROR: File not found: {file_path}")
        return pd.DataFrame() # Return empty if file not found
    except ValueError as ve: # Catch header detection error
        print(f"ERROR: Processing file {os.path.basename(file_path)} failed - {ve}")
        return pd.DataFrame()
    except Exception as e:
        print(f"ERROR: Unexpected error processing file {os.path.basename(file_path)}: {e}")
        import traceback
        traceback.print_exc() # Print full traceback for unexpected errors
        return pd.DataFrame()

# --- 6. 样式设置 ---
def classify_fields(df):
    basic_fields = [col for col in df.columns if "姓名" in col or "人员" in col or "部门" in col or "编号" in col or "身份证" in col or "职级" in col]
    stat_fields = [col for col in df.columns if "编制" in col or "身份" in col or "财政供养" in col or "统发" in col or "合计" in col or "小计" in col]
    income_fields = [col for col in df.columns if any(key in col for key in ["工资", "津贴", "补贴", "绩效", "奖金"])]
    deduct_fields = [col for col in df.columns if "扣" in col or "缴" in col or "个税" in col or "所得税" in col]

    classified = set()
    field_styles = {}
    for field in basic_fields:
        if field not in classified: field_styles[field] = "BASIC"; classified.add(field)
    for field in stat_fields:
        if field not in classified: field_styles[field] = "STAT"; classified.add(field)
    for field in income_fields:
        if field not in classified: field_styles[field] = "INCOME"; classified.add(field)
    for field in deduct_fields:
        if field not in classified: field_styles[field] = "DEDUCT"; classified.add(field)
    return field_styles

def get_column_width(cell):
    """计算单元格内容的适当列宽"""
    value = str(cell.value) if cell.value is not None else ""
    max_width = 0
    
    # 计算字符宽度
    for char in value:
        if '\u4e00' <= char <= '\u9fff':  # 中文字符
            max_width += 2.1
        elif char.isdigit() or char.isspace():  # 数字和空格
            max_width += 1.1
        else:  # 其他字符
            max_width += 1.3
    
    # 考虑字体样式的影响
    if cell.font.bold:
        max_width *= 1.2
    
    # 根据字体大小调整
    if cell.font.size:
        max_width *= cell.font.size / 11  # 11为默认字号
        
    return max_width

def format_excel_with_styles(filepath, output_path, year, month):
    wb = load_workbook(filepath)
    ws = wb.active

    title = f"{year}年{month:02d}月工资基金 机关工资发放表（实发）"
    date_str = datetime.today().strftime("制表时间：%Y 年 %m 月 %d 日")

    ws.insert_rows(1)
    ws.insert_rows(2)

    max_col = ws.max_column
    ws.merge_cells(f"A1:{get_column_letter(max_col)}1")
    ws["A1"] = title
    ws["A1"].font = Font(size=20, bold=True)
    ws["B2"] = "单位名称：高新区财政局"
    ws["B2"].font = Font(bold=True)
    ws["G2"] = date_str
    ws["G2"].font = Font(bold=True)

    headers = [cell.value for cell in ws[3]]
    style_map = {
        "BASIC": PatternFill("solid", fgColor="DCE6F1"),
        "STAT": PatternFill("solid", fgColor="EAEAEA"),
        "INCOME": PatternFill("solid", fgColor="E2F0D9"),
        "DEDUCT": PatternFill("solid", fgColor="FCE4D6"),
    }
    field_class = classify_fields(pd.DataFrame(columns=headers))

    for col_idx, col_name in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        style_key = field_class.get(col_name)
        if style_key in style_map:
            cell.fill = style_map[style_key]

    # 更新列宽度设置逻辑
    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_width = 0
        for cell in column_cells:
            width = get_column_width(cell)
            max_width = max(max_width, width)
        # 设置最小宽度并添加一些padding
        final_width = max(8, max_width + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(final_width, 50)  # 限制最大宽度为50

    for col_idx, col_name in enumerate(headers, 1):
        values = [ws.cell(row=row, column=col_idx).value for row in range(4, ws.max_row+1)]
        if all(v is None or str(v).strip() == "" for v in values):
            ws.column_dimensions[get_column_letter(col_idx)].hidden = True

    ws.freeze_panes = "H4"
    wb.save(output_path)
